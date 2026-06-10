from __future__ import annotations

"""
Exporta a aba principal do Banco NavalForge para CSV.

Uso:
    python scripts/export_xlsx_to_csv.py \
      --input NavalForge_Banco_Geral_Embarcacoes_4a20m_v119_teste_real_aplicado_1000_fichas.xlsx \
      --output data/navalforge_embarcacoes_semelhantes_r0.csv

Observação:
    Este script usa openpyxl apenas para leitura/exportação da planilha fonte.
"""

import argparse
import csv
import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_SHEET = "Fichas_Reais_v31_v100"


def export_xlsx_to_csv(input_path: Path, output_path: Path, sheet_name: str = DEFAULT_SHEET) -> dict:
    wb = load_workbook(input_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba não encontrada: {sheet_name}. Abas disponíveis: {wb.sheetnames}")

    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        if any(v is not None and str(v).strip() != "" for v in row):
            rows.append(list(row))

    if not rows:
        raise ValueError("Nenhuma linha encontrada na aba selecionada.")

    header = rows[0]
    last_col = max(i for i, v in enumerate(header) if v is not None and str(v).strip() != "") + 1
    rows = [r[:last_col] for r in rows]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerows(rows)

    data = output_path.read_bytes()
    manifest = {
        "arquivo": output_path.name,
        "fonte_xlsx": input_path.name,
        "aba": sheet_name,
        "linhas_total_incluindo_cabecalho": len(rows),
        "registros_dados": len(rows) - 1,
        "colunas": last_col,
        "tamanho_bytes": len(data),
        "sha256": hashlib.sha256(data).hexdigest(),
        "encoding": "utf-8-sig",
    }

    manifest_path = output_path.with_name(output_path.stem + "_manifest.json")
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return manifest


def main() -> None:
    parser = argparse.ArgumentParser(description="Exporta Banco NavalForge XLSX para CSV R0.")
    parser.add_argument("--input", required=True, help="Caminho da planilha .xlsx")
    parser.add_argument("--output", default="data/navalforge_embarcacoes_semelhantes_r0.csv", help="CSV de saída")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Nome da aba a exportar")
    args = parser.parse_args()

    manifest = export_xlsx_to_csv(Path(args.input), Path(args.output), args.sheet)
    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
